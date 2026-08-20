"""
Excel export endpoints — Phase 5.

GET /api/export/master               → P5-T5 master schedule
GET /api/export/department/{name}    → P5-T6 per-department schedule
GET /api/export/infeasibility        → P6-T6 infeasibility report

Both endpoints read the last solved schedule from the session store.
Both files are RTL-formatted, one-sheet workbooks (except infeasibility which has 4 sheets).
"""
from __future__ import annotations

import io
import datetime
from urllib.parse import quote

import pandas as pd
from fastapi import APIRouter, HTTPException, Path
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.session_store import get_session

router = APIRouter(prefix="/api/export", tags=["export"])

ALL_DEPTS = ["اتصالات", "برمجيات", "تقنية معلومات", "ذكاء اصطناعي", "شبكات الحاسوب", "صناعية", "طاقة متجددة", "ميكاترونيات"]

# ─── Matrix Grid Styles ──────────────────────────────────────────────────────────
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
DAY_FILL = PatternFill("solid", fgColor="FFFF00")  # Yellow
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
TOTALS_FONT = Font(bold=True, color="FF0000", name="Arial", size=11)
BOLD_FONT = Font(bold=True, name="Arial", size=11)
NORMAL_FONT = Font(name="Arial", size=11)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DAY_ALIGN = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)


def _build_matrix_sheet(ws, schedule_dict: dict, graph, target_depts: list[str]):
    ws.sheet_view.rightToLeft = True

    headers = ["الأيام", "المقررات"] + target_depts + ["الإجمالي الكلي"]
    ws.column_dimensions[get_column_letter(1)].width = 6
    ws.column_dimensions[get_column_letter(2)].width = 40
    for i in range(len(target_depts)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 12
    ws.column_dimensions[get_column_letter(3 + len(target_depts))].width = 14

    current_row = 1

    for iso_date, courses in sorted(schedule_dict.items()):
        day_courses = []
        for c in courses:
            cid = c["course_id"]
            info = graph.course_map.get(cid)
            if not info:
                continue
            
            course_variants = {}
            for d in target_depts:
                if d in info.variants:
                    course_variants[d] = info.variants[d]
                    
            if course_variants:
                day_courses.append((cid, info, course_variants))

        if not day_courses:
            continue

        start_row = current_row

        # Write Header Row
        for col_i, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_i)
            cell.value = header
            cell.font = BOLD_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
        
        ws.row_dimensions[current_row].height = 60
        current_row += 1

        # Write Totals Row
        dept_totals = {d: 0 for d in target_depts}
        day_grand_total = 0
        for (_, info, variants) in day_courses:
            for dept in target_depts:
                if dept in info.dept_students:
                    c_dept_count = len(info.dept_students[dept])
                    dept_totals[dept] += c_dept_count
                    day_grand_total += c_dept_count

        ws.cell(row=current_row, column=2, value="").border = THIN_BORDER
        for i, dept in enumerate(target_depts, start=3):
            cell = ws.cell(row=current_row, column=i)
            cell.value = str(dept_totals[dept]) if dept_totals[dept] > 0 else ""
            cell.font = TOTALS_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
        
        total_cell = ws.cell(row=current_row, column=3 + len(target_depts))
        total_cell.value = str(day_grand_total) if day_grand_total > 0 else ""
        total_cell.font = TOTALS_FONT
        total_cell.border = THIN_BORDER
        total_cell.alignment = CENTER_ALIGN
        
        current_row += 1

        # Write Course Rows
        for (cid, info, course_variants) in day_courses:
            names_list = []
            for dept, variants in course_variants.items():
                for v in variants:
                    level_str = f" {v[0]}" if v[0] else ""
                    names_list.append(f"{v[1]}{level_str}")
            
            unique_names = list(dict.fromkeys(names_list))
            names_str = "\n".join(unique_names)

            ws.cell(row=current_row, column=2, value=names_str).font = BOLD_FONT
            ws.cell(row=current_row, column=2).border = THIN_BORDER
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            
            course_grand_total = 0
            for i, dept in enumerate(target_depts, start=3):
                cell = ws.cell(row=current_row, column=i)
                count = len(info.dept_students.get(dept, set()))
                cell.value = count if count > 0 else ""
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN
                course_grand_total += count

            total_cell = ws.cell(row=current_row, column=3 + len(target_depts))
            total_cell.value = course_grand_total if course_grand_total > 0 else ""
            total_cell.font = BOLD_FONT
            total_cell.border = THIN_BORDER
            total_cell.alignment = CENTER_ALIGN

            # Expand row height to fit all variant names gracefully
            num_lines = max(1, len(unique_names))
            ws.row_dimensions[current_row].height = max(30, 20 * num_lines)

            current_row += 1

        # Format Date Cell
        dt = datetime.date.fromisoformat(iso_date)
        arabic_days = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الاحد"]
        arabic_months = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو", "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]
        day_name = arabic_days[dt.weekday()]
        month_name = arabic_months[dt.month - 1]
        
        def to_arabic_numeral(n: int) -> str:
            return "".join(chr(ord(c) + 1584) for c in str(n))
            
        date_str = f"{day_name} {to_arabic_numeral(dt.day)} {month_name} {to_arabic_numeral(dt.year)}"

        ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1)
        day_cell = ws.cell(row=start_row, column=1)
        day_cell.value = date_str
        day_cell.font = BOLD_FONT
        day_cell.fill = DAY_FILL
        day_cell.alignment = DAY_ALIGN

        for r in range(start_row, current_row):
            ws.cell(row=r, column=1).border = THIN_BORDER

        # empty row space between days
        current_row += 1


def _stream_wb(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# P5-T5 — Master schedule export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/master")
async def export_master(session_id: str):
    session = _get_valid_session(session_id)
    last = session.validation_report.get("last_schedule")
    if not last:
        raise HTTPException(status_code=422, detail="لم يتم تشغيل الجدولة بعد لهذه الجلسة.")

    wb = Workbook()
    ws = wb.active
    ws.title = "الجدول الكامل"
    _build_matrix_sheet(ws, last["schedule_dict"], last["graph"], ALL_DEPTS)
    
    return _stream_wb(wb, "الجدول_الكامل.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# P5-T6 — Per-department schedule export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/department/{dept_name}")
async def export_department(
    dept_name: str = Path(..., description="اسم القسم بالعربية"),
    session_id: str = "",
):
    session = _get_valid_session(session_id)
    last = session.validation_report.get("last_schedule")
    if not last:
        raise HTTPException(status_code=422, detail="لم يتم تشغيل الجدولة بعد لهذه الجلسة.")

    graph = last["graph"]
    
    # Verify department has courses
    has_courses = False
    for iso_date, courses in last["schedule_dict"].items():
        for c in courses:
            info = graph.course_map.get(c["course_id"])
            if info and dept_name in info.variants:
                has_courses = True
                break
        if has_courses:
            break
            
    if not has_courses:
        raise HTTPException(
            status_code=404,
            detail=f"القسم '{dept_name}' غير موجود في بيانات الجلسة أو ليس له مواد مجدولة.",
        )

    safe_name = dept_name.replace("/", "_").replace("\\", "_")
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"جدول {safe_name[:25]}"
    _build_matrix_sheet(ws, last["schedule_dict"], last["graph"], [dept_name])

    return _stream_wb(wb, f"جدول_{safe_name}.xlsx")


# ─── Helper ───────────────────────────────────────────────────────────────────

def _get_valid_session(session_id: str):
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id مطلوب.")
    session = get_session(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="الجلسة غير موجودة أو انتهت صلاحيتها.")
    return session

# ─────────────────────────────────────────────────────────────────────────────
# P6-T6 — Infeasibility report export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/infeasibility")
async def export_infeasibility(session_id: str):
    """
    Download the infeasibility report with four sheets:
    Summary, Top Students, Bottleneck Courses, Suggestions.
    """
    session = _get_valid_session(session_id)
    last = session.validation_report.get("last_schedule")
    if not last:
        raise HTTPException(status_code=422, detail="لم يتم تشغيل الجدولة بعد لهذه الجلسة.")

    result = last["result"]
    if getattr(result, "status", None) != "INFEASIBLE":
        raise HTTPException(status_code=400, detail="الجدول ليس في حالة غير ممكنة.")

    # Using the old styling definitions since this is a data report, not the matrix UI
    HEADER_FILL = PatternFill("solid", fgColor="1c2330")
    HEADER_FONT = Font(bold=True, color="00b4d8", name="Cairo", size=11)
    ACCENT_FILL = PatternFill("solid", fgColor="161b22")
    BODY_FONT   = Font(color="e6edf3", name="Cairo", size=10)

    def _style_ws_inf(ws, headers: list[str], col_widths: list[int]) -> None:
        ws.sheet_view.rightToLeft = True
        for col_i, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_i)
            cell.value = header
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_i)].width = width
        ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Sheet 1: Summary
        df_summary = pd.DataFrame([
            ("الأيام المتاحة", result.available_days),
            ("الحد الأدنى المطلوب", result.min_days_required),
            ("أيام إضافية مطلوبة", result.additional_days_needed),
        ], columns=["المقياس", "القيمة"])
        df_summary.to_excel(writer, index=False, sheet_name="الملخص")
        _style_ws_inf(writer.sheets["الملخص"], ["المقياس", "القيمة"], [30, 20])

        # Sheet 2: Top Students
        if result.top_students:
            rows = [(s["name"], s["course_count"], " · ".join(s["courses"])) for s in result.top_students]
            df_students = pd.DataFrame(rows, columns=["اسم الطالب", "عدد المواد", "المواد"])
            df_students.to_excel(writer, index=False, sheet_name="الطلاب الأكثر تأثيرا")
            _style_ws(writer.sheets["الطلاب الأكثر تأثيرا"], ["اسم الطالب", "عدد المواد", "المواد"], [40, 15, 60])

        # Sheet 3: Bottleneck Courses
        if result.bottleneck_courses:
            rows = [
                (c["course_id"], c["degree"], " | ".join(f"{d}: {', '.join(v['display_name'] for v in deptVs)}" for d, deptVs in c["variants"].items()))
                for c in result.bottleneck_courses
            ]
            df_courses = pd.DataFrame(rows, columns=["رمز المادة", "درجة التعارض", "الأقسام والمسميات"])
            df_courses.to_excel(writer, index=False, sheet_name="المواد الأكثر تعارضا")
            _style_ws(writer.sheets["المواد الأكثر تعارضا"], ["رمز المادة", "درجة التعارض", "الأقسام والمسميات"], [15, 15, 60])

        # Sheet 4: Suggestions
        if result.suggestions:
            rows = [(s["id"], s["action"], s["message"]) for s in result.suggestions]
            df_suggs = pd.DataFrame(rows, columns=["رقم", "الإجراء", "التفاصيل"])
            df_suggs.to_excel(writer, index=False, sheet_name="توصيات الحل")
            _style_ws(writer.sheets["توصيات الحل"], ["رقم", "الإجراء", "التفاصيل"], [10, 20, 80])

    return _stream(buf, "تقرير_عدم_إمكانية_الجدولة.xlsx")
